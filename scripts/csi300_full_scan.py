#!/usr/bin/env python3
"""
沪深300 缠论全量分析 — 完整版
生产XGBoost + V4.5 + GZK + 三维评分 + 风控计划 + PE + 宏观
输出: ~/chan_hs300_full_YYYYMMDD.xlsx
"""
import os, sys, csv, time, pickle, argparse
from datetime import datetime
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
REF = os.path.join(HERE, '..', 'references')
MODELS = os.path.join(HERE, '..', 'models')

from data import fetch_kline_a, fetch_a_quotes, load_a_stocks
sys.path.insert(0,os.path.join(HERE,'..','chanpy'))
from chan_engine_v5 import analyze as chan_analyze, get_bsp_label
print('[v5 engine]',end=' ')
from scorer import extract_features
from gann_enhance import compute_gann_enhance
from macro import load_macro, macro_signal
from sector_heat import sector_signal, get_sector_heat
from volume_sector import volume_analysis
from smc_insight import smc_analysis
from event_calendar import format_calendar
from enhanced_tools import (
    v45_experience_score, gzk_score, compute_3d_score, 
    check_risk, full_enhanced_analysis,
    mainline_score
)
from risk_engine import (atr_equal_risk_notional, risk_of_ruin,
                         GateContext, eval_all_gates, ExitPolicy, DSLTracker)
from cyq_chip import calc_cyq, calc_chip_score, fetch_turnover_batch
from sr_zones import sr_zones, split_sr

# 风险引擎用户持仓(用于 opposite_direction_guard 禁止金字塔加仓)
USER_POSITIONS = [
    {'coin': '601899', 'side': 'long'},  # 紫金矿业
    {'coin': '600362', 'side': 'long'},  # 江西铜业
    {'coin': '000630', 'side': 'long'},  # 铜陵有色
    {'coin': '300866', 'side': 'long'},  # 安克创新
]

# ═══════════════ Model Loading ═══════════════
_prod_model = None
_prod_meta = None

def load_models():
    global _prod_model, _prod_meta
    if _prod_model is not None:
        return _prod_model
    # V2: 使用不可变 bundle，取消 latest fallback；缺失即明确报错
    try:
        from model_bundle import load_bundle
        bundle = load_bundle()
        _prod_model = bundle['model']
        _prod_meta = bundle['schema']   # {'feature_names', 'feature_hash', 'n_features'}
        print(f'[bundle {bundle["model_id"]} {bundle["schema"]["n_features"]}维 '
              f'hash={bundle["schema"]["feature_hash"][:8]}]', end=' ')
    except Exception as e:
        print(f'[model ERROR] {e}', end=' ')
        _prod_model = None
        _prod_meta = None
    return _prod_model

def predict_score(feats, model, feat_order=None):
    """Predict score using a given XGBoost model"""
    if model is None:
        return None
    try:
        if feat_order is not None:
            # Production model: use explicit feature order
            vec = np.array([[feats.get(k, 0.0) for k in feat_order]])
        else:
            nf = model.n_features_in_
            vec = np.array([[feats[k] for k in sorted(feats.keys())[:nf]]])
        return int(model.predict_proba(vec)[0, 1] * 100)
    except:
        return None

# ═══════════════ CSI 300 Stock List ═══════════════
def load_csi300_stocks(asof_date=None):
    """Point-in-Time 沪深300成分股（含科创板688，不再按代码前缀排除）。"""
    try:
        from universe import get_csi300_members
        members = get_csi300_members(asof_date)
    except Exception:
        members = None
    if members:
        # 仅剔除 ST / 退市整理（eligibility 过滤，非板块排除）
        return [(c, n) for c, n in members if 'ST' not in n and '退' not in n]
    # 兜底：membership 数据缺失时回退旧静态快照（此时也不排除688）
    codes = []
    with open(os.path.join(REF, 'hs300_stocks.csv')) as f:
        for row in csv.DictReader(f):
            c = row['成分券代码'].strip()
            n = row['成分券名称'].strip().strip('"')
            if 'ST' in n or '退' in n:
                continue
            codes.append((c, n))
    return codes

# ═══════════════ PE Batch Query (baostock) ═══════════════
def fetch_pe_batch(stocks, cache_file=None):
    """Batch query peTTM for all stocks via baostock. Returns {code: pe}"""
    if cache_file and os.path.exists(cache_file):
        with open(cache_file, 'rb') as f:
            return pickle.load(f)
    
    import baostock as bs
    bs.login()
    pe_map = {}
    for idx, (code, name) in enumerate(stocks):
        if idx % 50 == 0:
            print(f'    PE查询: {idx}/{len(stocks)}', flush=True)
        try:
            sym = 'sh.' + code if code.startswith('6') else 'sz.' + code
            from datetime import timedelta
            end_dt = datetime.now() - timedelta(days=1)
            start_dt = end_dt - timedelta(days=5)
            rs = bs.query_history_k_data_plus(sym, 'date,peTTM',
                start_date=start_dt.strftime('%Y-%m-%d'),
                end_date=end_dt.strftime('%Y-%m-%d'),
                frequency='d')
            while rs.next():
                d = rs.get_row_data()
                if len(d) > 1 and d[1]:
                    pe_map[code] = round(float(d[1]), 2)
        except:
            pass
    bs.logout()
    
    if cache_file:
        os.makedirs(os.path.dirname(cache_file) or '.', exist_ok=True)
        with open(cache_file, 'wb') as f:
            pickle.dump(pe_map, f)
    
    return pe_map

# ═══════════════ 基本面六项批量查询 (baostock) ═══════════════
def _latest_quarters(n=6):
    """返回最近 n 个 (year, quarter) 列表, 从当前季度往回推。"""
    from datetime import datetime
    now = datetime.now()
    y, m = now.year, now.month
    q = (m - 1) // 3 + 1
    out = []
    for _ in range(n):
        out.append((y, q))
        q -= 1
        if q == 0:
            q = 4
            y -= 1
    return out


def fetch_fundamentals_batch(stocks, cache_file=None):
    """批量拉基本面六项(PE/PB/ROE/毛利率/负债率/净利润同比), baostock。
    返回 {code: {pe, pb, roe, gross_margin, debt_ratio, profit_yoy}}"""
    if cache_file and os.path.exists(cache_file):
        with open(cache_file, 'rb') as f:
            return pickle.load(f)
    import baostock as bs
    from datetime import timedelta
    bs.login()
    fund_map = {}
    quarters = _latest_quarters(6)
    for idx, (code, name) in enumerate(stocks):
        if idx % 50 == 0:
            print(f'    基本面查询: {idx}/{len(stocks)}', flush=True)
        sym = ('sh.' + code) if code.startswith('6') else ('sz.' + code)
        e = {'pe': None, 'pb': None, 'roe': None, 'gross_margin': None,
             'debt_ratio': None, 'profit_yoy': None}
        try:
            # PE/PB: query_history_k_data_plus (最新交易日)
            end_dt = datetime.now() - timedelta(days=1)
            start_dt = end_dt - timedelta(days=10)
            rs = bs.query_history_k_data_plus(sym, 'date,peTTM,pbMRQ',
                start_date=start_dt.strftime('%Y-%m-%d'),
                end_date=end_dt.strftime('%Y-%m-%d'), frequency='d')
            while rs.next():
                d = rs.get_row_data()
                if len(d) > 1 and d[1] and e['pe'] is None:
                    e['pe'] = float(d[1])
                if len(d) > 2 and d[2] and e['pb'] is None:
                    e['pb'] = float(d[2])
            # ROE/毛利率: query_profit_data (最新季度, 字段 roeAvg=3 gpMargin=5)
            for y, q in quarters:
                rs = bs.query_profit_data(code=sym, year=y, quarter=q)
                while rs.next():
                    d = rs.get_row_data()
                    if len(d) > 3 and d[3] and e['roe'] is None:
                        e['roe'] = float(d[3])
                    if len(d) > 5 and d[5] and e['gross_margin'] is None:
                        e['gross_margin'] = float(d[5])
                if e['roe'] is not None and e['gross_margin'] is not None:
                    break
            # 负债率: query_balance_data (liabilityToAsset=7)
            for y, q in quarters:
                rs = bs.query_balance_data(code=sym, year=y, quarter=q)
                while rs.next():
                    d = rs.get_row_data()
                    if len(d) > 7 and d[7] and e['debt_ratio'] is None:
                        e['debt_ratio'] = float(d[7])
                if e['debt_ratio'] is not None:
                    break
            # 净利润同比: query_growth_data (YOYNI=5)
            for y, q in quarters:
                rs = bs.query_growth_data(code=sym, year=y, quarter=q)
                while rs.next():
                    d = rs.get_row_data()
                    if len(d) > 5 and d[5] and e['profit_yoy'] is None:
                        e['profit_yoy'] = float(d[5])
                if e['profit_yoy'] is not None:
                    break
        except Exception:
            pass
        # baostock 返回小数(0.18=18%), 转百分数匹配 fundamental_score 评分
        for k in ('roe', 'gross_margin', 'debt_ratio', 'profit_yoy'):
            if e[k] is not None:
                e[k] = round(e[k] * 100, 2)
        fund_map[code] = e
    bs.logout()
    if cache_file:
        os.makedirs(os.path.dirname(cache_file) or '.', exist_ok=True)
        with open(cache_file, 'wb') as f:
            pickle.dump(fund_map, f)
    return fund_map

# ═══════════════ 主力资金流批量查询 (新浪 MoneyFlow) ═══════════════
def fetch_flow_batch(codes, sleep=0.08):
    """批量取近5日主力净流入累计(元)。返回 {code: flow_5d}"""
    import subprocess, json
    flows = {}
    for idx, code in enumerate(codes):
        if idx % 50 == 0:
            print(f'    资金流查询: {idx}/{len(codes)}', flush=True)
        code = str(code).zfill(6)
        prefix = ('sh' + code) if code.startswith('6') else ('sz' + code)
        url = ('https://vip.stock.finance.sina.com.cn/quotes_service/api/json_v2.php/'
               'MoneyFlow.ssl_qsfx_zjlrqs?page=1&num=5&sort=opendate&asc=0&daima=' + prefix)
        try:
            r = subprocess.run(['curl', '-s', '--max-time', '6', url],
                               stdout=subprocess.PIPE, timeout=8)
            raw = r.stdout.decode('utf-8', errors='replace')
            data = json.loads(raw)
            flow_5d = 0.0
            if isinstance(data, list):
                for d in data:
                    try:
                        flow_5d += float(d.get('r0_net') or 0)
                    except Exception:
                        pass
            flows[code] = flow_5d
        except Exception:
            flows[code] = 0.0
        if sleep:
            time.sleep(sleep)
    return flows

# ═══════════════ Main Scan ═══════════════
def main():
    parser = argparse.ArgumentParser(description='沪深300缠论全量分析-完整版')
    parser.add_argument('--min-score', type=int, default=35, help='3D综合评分最低阈值')
    parser.add_argument('--output', type=str, default=None, help='输出路径')
    args = parser.parse_args()

    today = datetime.now().strftime('%Y%m%d')
    out_path = args.output or os.path.expanduser(f'~/chan_hs300_full_{today}.xlsx')

    print(f'🔬 沪深300缠论全量分析 — 完整版 (17模块)')
    print(f'  输出: {out_path}')

    # Load model
    prod_model = load_models()
    prod_feat_order = []  # v5+talib模型用 list 拼接特征, 不用 feat_order
    print(f'  生产模型: {"✅ 已加载" if prod_model is not None else "❌ 未找到"}')

    # Load stocks
    stocks = load_csi300_stocks()
    print(f'  股票池: {len(stocks)} 只 (CSI 300)')

    # Macro
    try:
        macro = load_macro()
        ms = macro_signal(macro)
        print(f'  宏观: DXY={macro.get("DXY",{}).get("value","?")} | {ms.get("bias","?")}')
    except Exception as e:
        macro = {}
        print(f'  宏观: 获取失败 ({e})')

    # 市场情绪温度计 (market_sentiment) — 全局择时约束
    sentiment = None
    sentiment_mult = 1.0  # 情绪降仓乘数
    try:
        from market_sentiment import market_sentiment
        sentiment = market_sentiment(today)
        if 'error' in sentiment:
            print(f'  情绪: {sentiment["error"]}')
            sentiment = None
        else:
            lv = sentiment['level']
            if '冰点' in lv:
                sentiment_mult = 0.5
            elif '冷' in lv:
                sentiment_mult = 0.7
            elif '过热' in lv:
                sentiment_mult = 0.8
            print(f'  情绪: {lv} (涨停{sentiment["zt_n"]} 炸板率{sentiment["break_rate"]}% 晋级{sentiment["advance_rate"]}% 最高{sentiment["max_board"]}板) → 仓位乘数×{sentiment_mult}')
    except Exception as e:
        print(f'  情绪: 获取失败 ({e})')

    # Sector heat
    try:
        heat = get_sector_heat()
        print(f'  板块: {len(heat)} 个')
    except:
        heat = {}

    # PE batch query (cached)
    pe_cache = os.path.expanduser('~/chan_pe_cache.pkl')
    pe_map = fetch_pe_batch(stocks, cache_file=pe_cache)
    print(f'  PE: {len(pe_map)} 只有效')

    # 基本面六项 (PE/PB/ROE/毛利率/负债率/净利润同比, baostock, cached)
    try:
        fund_cache = os.path.expanduser('~/chan_fundamentals_cache.pkl')
        fund_map = fetch_fundamentals_batch(stocks, cache_file=fund_cache)
        n_fund = sum(1 for e in fund_map.values() if any(v is not None for v in e.values()))
        print(f'  基本面六项: {n_fund} 只有效')
    except Exception as e:
        fund_map = {}
        print(f'  基本面六项: 获取失败 ({e})')

    # Fetch quotes
    print(f'  获取行情...')
    quotes = fetch_a_quotes([(c, n) for c, n in stocks])
    print(f'  行情: {len(quotes)} 只')

    # Pre-fetch all K-lines via 新浪数据源(baostock已拉黑)
    from tencent_data import fetch_kline
    kline_cache = {}
    for idx, (code, name) in enumerate(stocks):
        try:
            dates, opens, highs, lows, closes, vols = fetch_kline(code)
            if len(dates) >= 100:
                kline_cache[code] = (dates, opens, closes, highs, lows, vols)
        except:
            pass
    print(f'  K线缓存: {len(kline_cache)} 只')

    # 换手率: 东财接口限流严重, 筹码分布用成交量归一化兜底(覆盖100%), 东财仅可选精度提升
    turnover_map = {}

    # Scan
    name_map = {c: n for c, n in stocks}
    results = []

    # ═══════ 因子预处理: 预收集特征 + winsorize 去极值 (吸收 factor_preprocess) ═══════
    # 对 280 只股票的 101 维特征做横截面 1%-99% 缩尾, 让 XGBoost 打分不被极端特征值扭曲。
    # 两阶段: 先收集全部特征向量 → 按列 1%-99% clip → 主循环复用 winsorized 特征。
    print(f'  [预处理] 预收集特征 + winsorize 去极值...')
    feat_cache = {}
    for _code, _name in stocks:
        if _code not in quotes:
            continue
        _data = kline_cache.get(_code)
        if not _data:
            continue
        _dates, _opens, _closes, _highs, _lows, _vols = _data
        try:
            _cur, _bsp_buy, _bsp_types, _px2, _zs_str, _pos = chan_analyze(
                _dates, _opens, _closes, _highs, _lows, _code)
        except Exception:
            continue
        _feats = extract_features(_closes, _highs, _lows, _opens, _vols, _bsp_buy, _bsp_types, _cur)
        try:
            from talib_features import add_talib_features, get_talib_feature_names
            _talib_feats = add_talib_features(_closes, _highs, _lows, _vols)
            if _prod_meta is not None:
                # V2: 用 bundle schema 的严格顺序构造特征（训练/推理同一 schema）
                _all = dict(_feats)
                _all.update(_talib_feats)
                _feat_vec = [_all.get(k, 0.0) for k in _prod_meta['feature_names']]
            else:
                _talib_names = get_talib_feature_names()
                _feat_vec = [_feats[k] for k in sorted(_feats.keys())] + \
                            [_talib_feats.get(k, 0.0) for k in _talib_names]
        except Exception:
            _feat_vec = [_feats[k] for k in sorted(_feats.keys())]
        feat_cache[_code] = (_feat_vec, _cur, _bsp_buy, _bsp_types, _zs_str, _pos,
                             _closes, _highs, _lows, _vols, _opens)

    # 横截面 1%-99% winsorize (每列独立裁剪极端值)
    if feat_cache:
        _codes = list(feat_cache.keys())
        _mat = np.array([feat_cache[c][0] for c in _codes], dtype=np.float64)
        for _j in range(_mat.shape[1]):
            _col = _mat[:, _j]
            _lo, _hi = np.nanpercentile(_col, [1, 99])
            _mat[:, _j] = np.clip(_col, _lo, _hi)
        for _i, _c in enumerate(_codes):
            _entry = feat_cache[_c]
            feat_cache[_c] = (_mat[_i].tolist(),) + _entry[1:]
    print(f'  [预处理] 特征收集 {len(feat_cache)} 只, winsorize 完成')

    # 关系特征 (DTW形态相似度/PageRank/中心性) — 全市场图结构
    rel_map = {}
    try:
        from relation_features import relation_features
        codes_list = list(feat_cache.keys())
        ret_matrix = []
        for c in codes_list:
            closes = feat_cache[c][6]  # closes 是 feat_cache 第6元素
            if len(closes) >= 60:
                rets = [closes[i] / closes[i-1] - 1 for i in range(len(closes)-60, len(closes))]
            else:
                rets = [0.0] * 60
            ret_matrix.append(rets)
        ret_matrix = np.array(ret_matrix)
        rel_feat = relation_features(ret_matrix)
        rel_map = {code: {'degree': float(rel_feat['degree_centrality'][i]),
                          'pagerank': float(rel_feat['pagerank'][i]),
                          'dtw_mean': float(rel_feat['dtw_similarity_mean'][i])}
                   for i, code in enumerate(codes_list)}
        print(f'  [关系特征] DTW/PageRank/中心性 完成 ({len(codes_list)}只)')
    except Exception as e:
        print(f'  [关系特征] 失败 ({e})')

    # 主力资金流批量拉取(近5日净流入)
    print('  [资金流] 批量拉取近5日主力净流入...')
    flow_map = fetch_flow_batch([c for c, _ in stocks], sleep=0.06)

    for idx, (code, name) in enumerate(stocks):
        if code not in quotes:
            continue
        if idx % 10 == 0:
            print(f'    {idx}/{len(stocks)}, found {len(results)}', flush=True)
        
        q = quotes[code]
        px = q['price']

        # 从预收集缓存取数据 (winsorized 特征 + chan 结果)
        cached = feat_cache.get(code)
        if not cached:
            continue
        feat_vec, cur, bsp_buy, bsp_types, zs_str, pos, closes, highs, lows, vols, opens = cached

        # XGBoost score — v5+talib 101维模型 (winsorized 特征)
        if prod_model is not None:
            try:
                nf = prod_model.n_features_in_
                vec = np.array([feat_vec[:nf]], dtype=np.float32)
                prod_xgb = int(prod_model.predict_proba(vec)[0, 1] * 100)
            except Exception:
                prod_xgb = None
        else:
            prod_xgb = None

        # 四框架增强 (江恩八分位 + MACD共振 + Ari动量)
        try:
            enh = compute_gann_enhance(np.array(closes), np.array(highs), np.array(lows))
            enhance_score = enh['enhance_score']
            enhance_detail = '{}|{}|{}'.format(
                enh['gann']['direction'], enh['macd']['resonance'], enh['ari']['env'])
        except Exception:
            enhance_score = 50.0
            enhance_detail = '-'

        # V4.5
        try:
            v45 = v45_experience_score(np.array(closes), np.array(highs), np.array(lows), np.array(vols))
            v45_score = v45.get('final_score', 0)
        except:
            v45_score = 0

        # GZK (returns dict with trend/deviation/k_ratio etc, NOT final_score)
        try:
            gzk = gzk_score(np.array(closes), np.array(highs), np.array(lows))
            # Compute GZK composite from available fields
            gzk_val = 0
            trend = gzk.get('trend', '')
            dev = gzk.get('deviation', '')
            k_ratio = gzk.get('k_ratio') or 0
            timing = gzk.get('timing_ok', False)
            # Trend scoring
            if trend == '走强': gzk_val += 30
            elif trend == '震荡': gzk_val += 20
            # Deviation scoring
            if dev == '超跌': gzk_val += 25
            elif dev == '正常': gzk_val += 15
            # K ratio scoring (capped)
            gzk_val += min(abs(k_ratio) * 15, 30)
            # Timing bonus
            if timing: gzk_val += 15
            gzk_val = max(0, min(100, round(gzk_val)))
        except:
            gzk_val = 0

        # 3D Composite — uses production XGB score as tech_score
        tech_score = prod_xgb if prod_xgb is not None else 0
        tech_score = max(0, min(100, tech_score))
        # fund 统一量纲: V4.5(0-30)归一化到0-100 + GZK(0-100), 各占50%
        v45_norm = min(100.0, v45_score / 30.0 * 100.0)
        fund_score = max(0, min(100, v45_norm * 0.5 + gzk_val * 0.5))
        # 3D: tech×0.55 + fund×0.45 (移除news硬编码, 原news=50浪费30%权重)
        score3d = compute_3d_score(tech_score, fund_score, 50,
                                   w_tech=0.55, w_fund=0.45, w_news=0.0)

        if score3d['composite'] < args.min_score:
            continue

        # PE from baostock batch query
        pe = pe_map.get(code)

        # 基本面六项评分 (fundamental_score) — 补"真基本面"维度
        fund_score6 = None
        fund_rating = '—'
        try:
            from fundamental_score import fundamental_score as fs6
            f = fund_map.get(code, {})
            if f and any(v is not None for v in f.values()):
                r = fs6(pe=(f.get('pe') or pe), pb=f.get('pb'), roe=f.get('roe'),
                        revenue_yoy=f.get('profit_yoy'), gross_margin=f.get('gross_margin'),
                        debt_to_asset=f.get('debt_ratio'))
                fund_score6 = r['total']   # 0-30
                fund_rating = r['rating']
        except Exception:
            pass

        # 关系特征 (pagerank 龙头识别 + dtw 形态相似度)
        rel_pagerank = rel_map.get(code, {}).get('pagerank') if rel_map else None
        rel_dtw = rel_map.get(code, {}).get('dtw_mean') if rel_map else None

        # YTD
        ytd = (closes[-1] / closes[-120] - 1) * 100 if len(closes) >= 120 else None

        # R:R and Entry/Stop/TP — calculate from 中枢
        label = get_bsp_label(bsp_buy, bsp_types, pos)
        
        # Parse zs for supports/resistances
        supports = []
        resistances = []
        if zs_str:
            for zs_part in zs_str.split(','):
                try:
                    zl, zh = map(float, zs_part.split('~'))
                    supports.append(zl)
                    resistances.append(zh)
                except:
                    pass

        # Calculate entry/stop/TP from最近的 中枢
        entry = None
        stop = None
        tp1 = None
        rr = None
        zl_low = None  # 中枢下沿(用于距买入区=现价距安全买点的距离)
        
        if supports and resistances:
            zl = zh = None
            # First: try to find the zhongshu the stock is currently inside
            for sl, sh in zip(supports, resistances):
                if sl <= px <= sh:
                    zl, zh = sl, sh
                    break
            
            if zl is None:
                # Not inside any zhongshu — find support (below) and resistance (above)
                # Track indices alongside values to avoid float equality bugs
                supports_below = [(i, s) for i, s in enumerate(supports) if s < px]
                resistances_above = [(i, r) for i, r in enumerate(resistances) if r > px]
                
                if supports_below and resistances_above:
                    # Between two zhongshu: use nearest below as support, nearest above as TP
                    best_s = max(supports_below, key=lambda x: x[1])  # closest support below
                    best_r = min(resistances_above, key=lambda x: x[1])  # closest resistance above
                    zl = best_s[1]
                    zh = best_r[1]
                elif supports_below:
                    # Below all zhongshu: use nearest support, paired resistance by same index
                    best_s = max(supports_below, key=lambda x: x[1])
                    zl = best_s[1]
                    zh = resistances[best_s[0]]
                elif resistances_above:
                    # Above all zhongshu: use nearest resistance, paired support by same index
                    best_r = min(resistances_above, key=lambda x: x[1])
                    zh = best_r[1]
                    zl = supports[best_r[0]]
                else:
                    # Fallback to nearest zhongshu center
                    nearest_idx = 0
                    min_dist = float('inf')
                    for i, (sl, sh) in enumerate(zip(supports, resistances)):
                        center = (sl + sh) / 2
                        dist = abs(px - center)
                        if dist < min_dist:
                            min_dist = dist
                            nearest_idx = i
                    zl = supports[nearest_idx]
                    zh = resistances[nearest_idx]
            
            if zl is not None:
                zl_low = zl  # 保存中枢下沿(距买入区锚)
            if zl is not None and zh is not None and bsp_buy:
                # 统一 R:R 逻辑: entry=现价, stop=ATR自适应(最少3%保护), tp1=结构目标
                # ATR14(波动率自适应止损, 替代固定3%)
                atr14 = 0.0
                if len(closes) >= 15:
                    trs = []
                    for i in range(1, len(closes)):
                        trs.append(max(highs[i] - lows[i],
                                       abs(highs[i] - closes[i-1]),
                                       abs(lows[i] - closes[i-1])))
                    atr14 = sum(trs[-14:]) / 14
                entry = round(px, 2)
                stop = round(px - max(2.0 * atr14, 0.03 * px), 2)
                if px < zh:
                    # 中枢内/中枢下: 目标=中枢上沿(阻力)
                    tp1 = round(zh, 2)
                else:
                    # 三买(已突破中枢上沿): 目标=投影(上沿+带宽)
                    tp1 = max(round(zh + (zh - zl), 2), round(entry * 1.10, 2))
                if entry > stop and stop > 0 and tp1 > entry:
                    rr = round((tp1 - entry) / (entry - stop), 1)
            elif 'Sell' in label:
                entry = px
                # For Sell: find support below (TP) and resistance above (stop)
                sell_zl = [s for s in supports if s < px]
                sell_zh = [r for r in resistances if r > px]
                if sell_zh:
                    stop = round(min(sell_zh) * 1.03, 2)
                else:
                    stop = round(px * 1.03, 2)
                if sell_zl:
                    tp1 = round(max(sell_zl), 2)  # nearest support below = TP
                else:
                    tp1 = round(px * 0.95, 2)
                if stop > entry and stop > 0 and tp1 < entry:
                    rr = round((entry - tp1) / (stop - entry), 1)
            # 中枢内等信号(非买卖) — 给出入场参考
            elif zl is not None and zh is not None and zl <= px <= zh:
                entry = round(zl + (zh - zl) * 0.1, 2)
                stop = round(zl * 0.97, 2)
                tp1 = round(zh, 2)
                if entry > stop and stop > 0:
                    rr = round((tp1 - entry) / (entry - stop), 1)
        
        rr = max(0, min(rr or 0, 20))  # cap at 20
        
        signal_type = 'buy' if bsp_buy else ('sell' if 'Sell' in label else 'neutral')

        # Determine if in central zone
        in_zs = False
        if zs_str and supports and resistances:
            for zl, zh in zip(supports, resistances):
                if zl <= px <= zh:
                    in_zs = True
                    break

        # Sector
        sec = sector_signal(code, heat=heat)
        
        # Volume
        vol_analysis = volume_analysis([float(x) for x in closes], [float(x) for x in vols])

        # Wyckoff state from volume + price context
        vol_sig = vol_analysis.get('signal', '')
        price5 = (closes[-1] / closes[-6] - 1) * 100 if len(closes) >= 6 else 0
        price20 = (closes[-1] / closes[-21] - 1) * 100 if len(closes) >= 21 else 0
        
        wyckoff = ''
        if in_zs and ('缩' in str(vol_sig) or '量价正常' in str(vol_sig)):
            wyckoff = 'S(弹簧/缩量区)' if price20 < 0 else 'A(吸筹待确认)'
        elif in_zs and '放' in str(vol_sig):
            wyckoff = 'B(放量突破)' if price5 > 0 else 'D(派发警告)'
        elif not in_zs and '放' in str(vol_sig) and price5 > 3:
            wyckoff = 'C(放量离开)'
        elif '顶背' in str(vol_sig):
            wyckoff = 'D(派发信号)'
        elif '底背' in str(vol_sig):
            wyckoff = 'A(吸筹底背离)'
        elif '背' in str(vol_sig):
            wyckoff = 'W(量价背离)'
        else:
            wyckoff = '-'

        # Tag
        tag = ''
        if in_zs and bsp_buy:
            tag = '中枢内买'
        elif in_zs and not bsp_buy:
            tag = '中枢内等信号'
        elif in_zs and ('Sell' in label):
            tag = '中枢内Sell'
        else:
            # Distance-based tag
            if supports:
                nearest_zl = min(supports, key=lambda z: abs(px - z))
                dist_pct = round((px - nearest_zl) / nearest_zl * 100, 0)
                tag = f'中枢外买(距{dist_pct:+.0f}%)'

        # Risk check
        blocked, risk_reasons = check_risk(code, name)

        # ══════ 风险引擎集成 (risk_engine) — ATR等风险仓位 + 门控 + DSL ══════
        risk_size = '—'   # 建议股数(100股整手)
        risk_gate = '-'    # 门控状态 PASS/BLOCK
        risk_ror = None    # 爆仓概率
        if bsp_buy and isinstance(entry, (int, float)) and isinstance(stop, (int, float)) \
                and entry > stop > 0:
            stop_dist = (entry - stop) / entry
            size = atr_equal_risk_notional(
                equity=1_000_000, risk_per_trade_pct=0.02,
                atr_abs=entry - stop, entry_px=entry, sl_atr_mult=1.0,
                max_trade_notional_usd=200_000, config_max_leverage=1)
            risk_size = int(size.notional_usd / entry / 100) * 100
            ctx = GateContext(
                confidence=score3d['composite'] / 100.0,
                current_positions=USER_POSITIONS,
                trade_notional_usd=size.notional_usd,
                daily_pnl=0.0,
                market_volume_24h_usd=1e8,
                coin=code, trade_side='long',
                has_binary_news_risk=False,
                equity=1_000_000, total_open_notional=0.0,
                composite_score=score3d['composite'])
            gates = eval_all_gates(ctx, {
                'min_confidence': 0.45, 'max_concurrent': 5,
                'max_trade_notional_usd': 200_000, 'max_daily_loss_usd': -30_000,
                'max_total_notional_pct': 1.0,
            }, regime='neutral')
            risk_gate = 'PASS' if not gates['blocked'] else 'BLOCK'
            win_rate = min(0.6, 0.4 + score3d['composite'] / 300)
            payoff = max(1.0, min(4.0, float(rr or 2.0)))
            risk_ror = risk_of_ruin(win_rate=win_rate, payoff_ratio=payoff,
                                    risk_per_trade_pct=0.02)

        # ══════ 筹码分布(CYQ) + 筹码评分 ══════
        chip_score_val = None
        chip_veto = False
        chip_detail = '—'
        chip_benefit = None
        chip_avg_cost = None
        if len(closes) >= 30:
            try:
                # 换手率: 东财取到用真实值, 否则用成交量归一化代理(默认均值3%)
                turnover_avg = turnover_map.get(code, 3.0)
                mean_vol = sum(vols) / len(vols) if vols else 1.0
                turnovers = [max(0.1, turnover_avg * (v / mean_vol)) for v in vols]
                cyq = calc_cyq(opens, closes, highs, lows, turnovers,
                               crange=120, cyq_days=210)
                if cyq.get('benefit_part') is not None:
                    chip_score_val, chip_veto, chip_detail = calc_chip_score(
                        cyq['benefit_part'], cyq['concentration_90'], px, cyq['avg_cost'])
                    chip_benefit = cyq['benefit_part']
                    chip_avg_cost = cyq['avg_cost']
            except Exception:
                pass

        # ══════ S&R 支撑阻力带 (sr_zones 聚类算法) ══════
        sr_score = 50.0
        sr_r = None   # 最近阻力
        sr_s = None   # 最近支撑
        sr_detail = '—'
        try:
            sr_zones_list = sr_zones(highs, lows, tolerance=0.005, top=8)
            sr_res, sr_sup = split_sr(sr_zones_list, px)
            sr_r = sr_res[0]['price'] if sr_res else None
            sr_s = sr_sup[0]['price'] if sr_sup else None
            if sr_r is not None and sr_s is not None and sr_r > sr_s:
                sr_pos = (px - sr_s) / (sr_r - sr_s)  # 0=贴支撑, 1=贴阻力
                sr_score = max(0.0, min(100.0, 100.0 - sr_pos * 100.0))
            elif sr_s is not None and sr_s > 0:
                # 只有支撑: 现价距支撑越近越安全(贴支撑=100, 距支撑>25%→0)
                dist_s = (px - sr_s) / sr_s * 100
                sr_score = max(0.0, min(100.0, 100.0 - dist_s * 4))
            elif sr_r is not None and sr_r > 0:
                # 只有阻力: 现价距阻力越远越安全(贴阻力=0, 距阻力>25%→100)
                dist_r = (sr_r - px) / sr_r * 100
                sr_score = max(0.0, min(100.0, dist_r * 4))
            if sr_r is not None or sr_s is not None:
                sr_detail = f'R:{sr_r if sr_r else "-"}/S:{sr_s if sr_s else "-"}'
        except Exception:
            pass

        # ══════ 三锚共振度: 中枢内 + S&R贴支撑 + 筹码健康 ══════
        resonance = 0
        if in_zs:
            resonance += 1  # 锚1: 中枢内买
        if sr_score >= 70:
            resonance += 1  # 锚2: S&R贴支撑(安全买点)
        if chip_score_val is not None and chip_score_val >= 60:
            resonance += 1  # 锚3: 筹码健康
        resonance_label = {3: '三重共振', 2: '双锚', 1: '单锚', 0: '无锚'}[resonance]

        # 主力资金流评分: 近5日净流入(亿元) → 0-100 (每+1亿→+50分)
        flow_5d = flow_map.get(code, 0.0)
        flow_score = max(0.0, min(100.0, 50.0 + flow_5d / 1e8 * 50.0))

        results.append({
            'code': code,
            'name': name,
            'price': px,
            'pe': pe,
            'ytd': round(ytd, 1) if ytd is not None else None,
            'prod_xgb': prod_xgb if prod_xgb is not None else 0,
            'score3d': score3d['composite'],
            'grade': score3d['grade'],
            'position_pct': score3d['position'] * 100,
            'enhance': enhance_score,
            'enhance_detail': enhance_detail,
            'rr': round(rr, 1),
            'zs': zs_str or '-',
            'zl': round(zl_low, 2) if zl_low is not None else None,
            'in_zs': '是' if in_zs else '否',
            'bsp': label,
            'v45': v45_score,
            'gzk': round(gzk_val, 1),
            'entry': round(entry, 2) if entry is not None else '—',
            'stop': round(stop, 2) if stop is not None else '—',
            'tp1': round(tp1, 2) if tp1 is not None else '—',
            'risk_status': 'BLOCKED' if blocked else 'OK',
            'tag': tag,
            'vol_signal': vol_analysis.get('signal', '-'),
            'wyckoff': wyckoff,
            'sector': sec.get('sector', '-'),
            'risk_size': risk_size,
            'risk_gate': risk_gate,
            'risk_ror': round(risk_ror, 4) if risk_ror is not None else '—',
            'chip_score': chip_score_val if chip_score_val is not None else '—',
            'chip_veto': chip_veto,
            'chip_benefit': chip_benefit,
            'chip_avg_cost': chip_avg_cost,
            'sr_score': round(sr_score, 1),
            'sr_detail': sr_detail,
            'sr_r': sr_r,
            'sr_s': sr_s,
            'resonance': resonance,
            'resonance_label': resonance_label,
            'flow_5d': flow_5d,
            'flow_score': round(flow_score, 1),
            'fund_score6': fund_score6,
            'fund_rating': fund_rating,
            'rel_pagerank': rel_pagerank,
            'rel_dtw': rel_dtw,
        })

    print(f'  完成: {len(results)} 只信号')

    # Sort by 3D score
    results.sort(key=lambda x: -x['score3d'])

    # ═══════════════ Excel Output ═══════════════
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Signals ---
    ws = wb.active
    ws.title = f'{today}信号'

    headers = ['代码', '名称', '现价', 'PE', 'YTD%', '生产XGB', '3D分', 
               '等级', '仓位%', 'R:R', '中枢', '中枢内', 'BSP', 'V4.5', 'GZK',
               '买入', '止损', 'TP1', '风控', '标签', '威科夫',
               '风险仓位', '风险门控', '爆仓概率', 'S&R带', 'S&R分', '共振', '资金(亿)',
               '基本面', '龙头']

    # Header style
    hdr_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    hdr_font = Font(color='ffffff', bold=True, size=10)
    green_fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
    yellow_fill = PatternFill(start_color='fff8e1', end_color='fff8e1', fill_type='solid')
    red_fill = PatternFill(start_color='fce4ec', end_color='fce4ec', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='cccccc'),
        right=Side(style='thin', color='cccccc'),
        top=Side(style='thin', color='cccccc'),
        bottom=Side(style='thin', color='cccccc')
    )

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, r in enumerate(results):
        row = i + 2
        vals = [
            r['code'], r['name'], r['price'], r['pe'], r['ytd'],
            r['prod_xgb'], r['score3d'],
            r['grade'], r['position_pct'], r['rr'],
            r['zs'], r['in_zs'], r['bsp'],
            r['v45'], r['gzk'],
            r['entry'], r['stop'], r['tp1'],
            r['risk_status'], r['tag'], r['wyckoff'],
            r['risk_size'], r['risk_gate'], r['risk_ror'],
            r['sr_detail'], r['sr_score'], r['resonance_label'],
            round(r['flow_5d'] / 1e8, 2),
            (f"{r['fund_score6']}/{r['fund_rating']}" if r.get('fund_score6') is not None else '—'),
            ('龙头' if (r.get('rel_pagerank') or 0) > 0.004 else '-')
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color coding
        if r['grade'] == 'A':
            for c in range(1, len(headers)+1):
                ws.cell(row, c).fill = green_fill
        elif r['risk_status'] == 'BLOCKED':
            for c in range(1, len(headers)+1):
                ws.cell(row, c).fill = red_fill
        elif '中枢内买' == r['tag']:
            for c in range(1, len(headers)+1):
                ws.cell(row, c).fill = green_fill

    # Column widths (24 columns: A-X)
    widths = {'A':8, 'B':12, 'C':8, 'D':6, 'E':7, 'F':9, 'G':5,
              'H':6, 'I':5, 'J':5, 'K':20, 'L':6, 'M':16, 'N':5, 'O':5,
              'P':8, 'Q':8, 'R':8, 'S':6, 'T':6, 'U':18,
              'V':9, 'W':9, 'X':9, 'Y':11, 'Z':6, 'AA':9, 'AB':9, 'AC':11, 'AD':6}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:AD{len(results)+1}'

    # --- Sheet 2: Macro (detailed) ---
    ws2 = wb.create_sheet('宏观')
    
    # Stats
    buys = sum(1 for r in results if 'Buy' in r['bsp'])
    sells = sum(1 for r in results if 'Sell' in r['bsp'])
    holds = sum(1 for r in results if 'Hold' in r['bsp'])
    inzs = sum(1 for r in results if r['in_zs'] == '是')
    grade_a = sum(1 for r in results if r['grade'] == 'A')
    grade_b = sum(1 for r in results if r['grade'] == 'B')
    buys_zs = sum(1 for r in results if r['in_zs'] == '是' and 'Buy' in r['bsp'])
    
    macro_lines = [
        [f'{today} 收盘 | 缠论+生产XGBoost+三维评分+风控 | 全功能扫描'],
        [''],
        ['━━━ 市场指数 ━━━'],
        [f'DXY: {macro.get("DXY",{}).get("value","?")} | US10Y: {macro.get("UST10Y",{}).get("value","?")}% | USD/CNY: {macro.get("USDCNY",{}).get("value","?")}'],
        [f'宏观方向: {macro_signal(macro).get("bias","中性")}'],
        [''],
        ['━━━ 市场情绪 ━━━'],
        [f'情绪: {sentiment["level"]} | 涨停{sentiment["zt_n"]} 炸板{sentiment["zb_n"]} 跌停{sentiment["dt_n"]} 炸板率{sentiment["break_rate"]}%' if sentiment else '情绪: 未获取(非交易日或接口不可达)'],
        [f'连板: 最高{sentiment["max_board"]}板 | 昨涨停溢价{sentiment["premium"]:+.2f}% | 晋级率{sentiment["advance_rate"]}% | 大面{sentiment["big_loss"]}只' if sentiment else ''],
        [f'仓位乘数: ×{sentiment_mult} | {sentiment["advice"]}' if sentiment else f'仓位乘数: ×{sentiment_mult}(默认)'],
        [''],
        ['━━━ 扫描统计 ━━━'],
        [f'CSI300成分股: {len(stocks)}只 | 有效信号: {len(results)}只'],
        [f'Buy: {buys} | Sell: {sells} | Hold: {holds}'],
        [f'中枢内买: {buys_zs} | 中枢内总计: {inzs}'],
        [''],
        ['━━━ 3D评分分布 ━━━'],
        [f'A级(≥65): {grade_a} | B级(≥50): {grade_b} | C级(≥35): {len(results)-grade_a-grade_b}'],
        [f'分数范围: {results[-1]["score3d"]:.0f}-{results[0]["score3d"]:.0f}'],
        [''],
        ['━━━ 综合判断 ━━━'],
        ['评分按三维(技术40%+基本30%+消息30%)加权，中枢内买点优先。'],
        ['A级标的可重仓(50%)，B级可配置(20-30%)，C级观望(10%)。'],
        ['仅中枢内买点享有完整安全边际，中枢外买需次级别确认。'],
    ]
    
    for row_idx, line in enumerate(macro_lines, 1):
        for col_idx, val in enumerate(line, 1):
            c = ws2.cell(row_idx, col_idx, val)
            if '━' in str(val): c.font = Font(bold=True, size=11)
    ws2.column_dimensions['A'].width = 80
    
    # --- Sheet 3: 综合推荐 ---
    ws3 = wb.create_sheet('综合推荐')
    rec_headers = ['#', '代码', '名称', '现价', '生产XGB', '3D分', '等级', '仓位%', 
                   'R:R', '四框架', '筹码', 'S&R带', 'V4.5', 'GZK', '中枢', '买入', '止损', 'TP1', '方向', '威科夫', '逻辑', '分类']
    for c, h in enumerate(rec_headers, 1):
        cell = ws3.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Rank 优化: XGB×0.35(生产模型第一) + R:R×0.15(盈亏比第二) + 3D×0.15 + 筹码×0.10 + 四框架×0.10 + 距买入区×0.10 + V4.5奖励
    # XGB是生产模型(AUC0.667, 统计验证), 权重最高; R:R盈亏比排第二
    for r in results:
        chip = r['chip_score'] if isinstance(r['chip_score'], (int, float)) else 50.0
        # R:R 期望值修正: 裸R:R × XGB胜率修正, 封顶6(深跌低位股R:R虚高, 低胜率压制)
        rr_val = r['rr'] if isinstance(r['rr'], (int, float)) and r['rr'] > 0 else 0
        xgb_adj = max(0.4, min(1.5, r['prod_xgb'] / 50.0))  # XGB50=1.0, 20=0.4, 75=1.5
        rr_expected = min(rr_val * xgb_adj, 6.0)  # 期望R:R封顶6
        rr_score = rr_expected * 16.67  # 期望6→100, 期望3→50, 期望1.5→25
        # 距买入区=现价距中枢下沿(安全买点): 双边惩罚, 贴着下沿→100分, 超跌或追高→降分
        dist_score = 50.0
        zl_v = r.get('zl'); price_v = r['price']
        if isinstance(zl_v, (int, float)) and isinstance(price_v, (int, float)) and zl_v > 0:
            dist_pct = (price_v - zl_v) / zl_v * 100  # 正=追高(下沿上方), 负=超跌(下沿下方)
            dist_score = max(0.0, min(100.0, 100.0 - abs(dist_pct) * 5.0))
        # S&R 评分: 现价在支撑带(安全)与阻力带(追高)之间的位置, 贴支撑=100
        sr_sc = r.get('sr_score') if isinstance(r.get('sr_score'), (int, float)) else 50.0
        resonance = r.get('resonance', 0) or 0
        flow_sc = r.get('flow_score') if isinstance(r.get('flow_score'), (int, float)) else 50.0
        # 基本面六项(30分制→100分制): 补真基本面维度
        fund6_sc = 50.0
        if isinstance(r.get('fund_score6'), (int, float)):
            fund6_sc = min(100.0, r['fund_score6'] / 30.0 * 100.0)
        # 龙头奖励: pagerank 高于平均(图核心节点)加分
        leader_bonus = 0
        pr = r.get('rel_pagerank')
        if isinstance(pr, (int, float)) and pr > 0.004:
            leader_bonus = 3
        r['_rank'] = (r['prod_xgb'] * 0.35          # XGB第一(生产模型AUC0.667)
                      + rr_score * 0.15             # R:R第二(期望值修正)
                      + r['score3d'] * 0.15         # 3D
                      + chip * 0.10                 # 筹码
                      + r['enhance'] * 0.05         # 四框架
                      + flow_sc * 0.05              # 主力资金流(近5日净流入)
                      + fund6_sc * 0.05             # 基本面六项(PE/PB/ROE/毛利率/负债率/增速)
                      + dist_score * 0.05           # 距买入区(中枢下沿锚)
                      + sr_sc * 0.05                # S&R支撑阻力带(贴支撑加分)
                      + leader_bonus                # 龙头奖励(pagerank图核心)
                      + (8 if resonance >= 3 else (3 if resonance == 2 else 0))  # 三锚共振奖励
                      + (10 if r['v45'] >= 8 else 0))
    results.sort(key=lambda x: -x['_rank'])

    # 阿娇否决项: 方向非多 / 威科夫派发(D) / 筹码否决 / R:R<1.5 / 距买入区>20%(不可操作)
    def veto(r):
        direction = '多' if 'Buy' in str(r.get('bsp', '')) else ('空' if 'Sell' in str(r.get('bsp', '')) else '观望')
        if direction != '多':
            return True
        wy = str(r.get('wyckoff', '') or '')
        if wy.startswith('D('):  # 派发信号
            return True
        if r.get('chip_veto'):
            return True
        # R:R<1.5 否决(盈亏比不划算, 止损空间>盈利空间1.5倍以上)
        rr_v = r.get('rr')
        if isinstance(rr_v, (int, float)) and rr_v < 1.5:
            return True
        # 现价距中枢下沿>20%(超跌或追高)否决
        zl_v = r.get('zl'); price_v = r.get('price')
        if isinstance(zl_v, (int, float)) and isinstance(price_v, (int, float)) and zl_v > 0:
            if abs((price_v - zl_v) / zl_v * 100) > 20:
                return True
        return False

    recs = [r for r in results if not veto(r)
            and r.get('in_zs') == '是'          # 强制中枢内(修复中枢外标的混入)
            and r.get('risk_gate') != 'BLOCK']  # 剔除风险门控BLOCK
    recs = recs[:15]
    for i, r in enumerate(recs):
        logic = '+'.join([x for x in [
            ('中枢内' if r['in_zs'] == '是' else ''),
            ('3D' + r['grade']) if r['grade'] in ('A','B') else '',
        ] if x])
        rr_s = r['rr'] if r['rr'] > 0 else '—'
        direction = '多' if 'Buy' in str(r.get('bsp','')) else ('空' if 'Sell' in str(r.get('bsp','')) else '观望')
        # 二维分类: 共振度 × XGB胜率 (位置安全 vs 会涨)
        high_xgb = (r['prod_xgb'] or 0) >= 47
        triple = (r.get('resonance') or 0) >= 3
        category = ('核心买入' if high_xgb and triple else
                    '进攻候选' if high_xgb else
                    '防御观察' if triple else '观望')
        vals = [i+1, r['code'], r['name'], r['price'], r['prod_xgb'],
                r['score3d'], r['grade'], r['position_pct'], rr_s,
                r['enhance'], r['chip_score'], r['sr_detail'], r['v45'], r['gzk'], r['zs'][:15],
                r['entry'], r['stop'], r['tp1'], direction, r['wyckoff'], logic, category]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(i+2, c, v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if i < 5: cell.fill = green_fill
    
    # Column widths for 综合推荐 (20 columns)
    for c, w in enumerate([4,8,10,7,9,4,6,4,5,7,6,11,5,5,15,7,7,7,5,8,18,10], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws3.freeze_panes = 'A2'


    
    # Print top 20
    print(f'\n{"="*60}')
    print(f'🏆 Top 20 (3D综合评分)')
    print(f'{"="*60}')
    print(f'{"排名":<4} {"名称":<12} {"现价":<8} {"YTD%":<7} {"生产XGB":<8} {"3D分":<6} {"等级":<4} {"BSP":<18} {"标签":<20} {"R:R":<5}')
    print('-' * 80)
    for i, r in enumerate(results[:20]):
        print(f'{i+1:<4} {r["name"]:<12} {r["price"]:<8} {str(r["ytd"]):<7} '
              f'{r["prod_xgb"]:<8} {r["score3d"]:<6} {r["grade"]:<4} '
              f'{r["bsp"]:<18} {r["tag"]:<20} {r["rr"]:<5}')

    print(f'\nSaved: {out_path} | {len(results)} signals')
    print(f'A级: {grade_a} | B级: {grade_b} | 中枢内买: {inzs}')

    # --- 组合优化(Alpha-Risk Blended HRP) ---
    try:
        from portfolio_optimize import quality_filter, alpha_blended_hrp
        
        # Stage 1: Quality filter
        qualified = quality_filter(results)
        print(f'\n[组合优化] 质量过滤: {len(qualified)}/{len(results)}只通过 ' +
              f'(Buy+中枢内+R:R≥1.5+生产XGB≥40)')
        
        if len(qualified) >= 3:
            # Stage 2-5: Alpha-Risk Blended HRP
            weights, metrics = alpha_blended_hrp(
                qualified, 
                tilt_factor=0.5, 
                max_position=0.20,
                min_position=0.02,
                max_stocks=30
            )
            
            if weights:
                ws4 = wb.create_sheet('组合优化')
                opt_headers = ['#', '代码', '名称', '权重%', 'Alpha', 'HRP基%', '倾斜', 
                               '现价', '生产XGB', '3D分', 'R:R', 'V4.5', '威科夫', '方向']
                for c, h in enumerate(opt_headers, 1):
                    cell = ws4.cell(1, c, h)
                    cell.fill = hdr_fill; cell.font = hdr_font
                    cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

                # Build lookup maps
                name_map = {r['code']: r['name'] for r in results}
                px_map = {r['code']: r['price'] for r in results}
                xgb_map = {r['code']: r['prod_xgb'] for r in results}
                s3d_map = {r['code']: r['score3d'] for r in results}
                rr_map = {r['code']: r['rr'] for r in results}
                v45_map = {r['code']: r['v45'] for r in results}
                wyckoff_map = {r['code']: r.get('wyckoff','-') for r in results}

                # 情绪降仓乘数: 冷/冰点降仓, 过热减仓
                if sentiment_mult != 1.0:
                    weights = {c: w * sentiment_mult for c, w in weights.items()}
                    lv = sentiment['level'] if sentiment else '未知'
                    print(f'  [情绪降仓] 权重×{sentiment_mult} (情绪{lv})')

                ranked = sorted(weights.items(), key=lambda x: -x[1])
                for i, (code, wt) in enumerate(ranked):
                    m = metrics.get(code, {})
                    name = name_map.get(code, code)
                    vals = [
                        i+1, code, name, round(wt*100, 1),
                        m.get('alpha', '-'), m.get('hrp_w', '-'), m.get('tilt', '-'),
                        px_map.get(code, '-'), xgb_map.get(code, '-'),
                        s3d_map.get(code, '-'), rr_map.get(code, '-'),
                        v45_map.get(code, '-'), wyckoff_map.get(code, '-'),
                        '多'
                    ]
                    for c, v in enumerate(vals, 1):
                        cell = ws4.cell(i+2, c, v)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')
                        if i < 3:
                            cell.font = Font(color='2F5496', bold=True)
                
                # Column widths
                for c, w in enumerate([4,8,10,7,6,7,6,6,8,6,5,5,14], 1):
                    ws4.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
                ws4.freeze_panes = 'A2'
    except Exception as e:
        print(f'  组合优化跳过: {e}')

    wb.save(out_path)


if __name__ == '__main__':
    main()
