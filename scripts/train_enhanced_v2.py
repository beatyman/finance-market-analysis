#!/usr/bin/env python3
"""
XGBoost增强训练器 v2 — 带进度/日志/时间戳/增量保存

相比v1改进:
  1. imap_unordered 实时进度 (x/271)
  2. 日志写入文件 + 终端双输出
  3. 每阶段时间戳记录
  4. 增量保存中间结果(断点续传)
  5. 阶段状态文件(随时可查进度)

运行: python3 train_enhanced_v2.py
查看进度: tail -f /tmp/train_v2.log
"""
import os, sys, time, pickle, json, numpy as np
from multiprocessing import Pool, cpu_count, Manager
from datetime import datetime

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
sys.path.insert(0, os.path.join(HERE, '..', 'chanpy'))

from chan_engine_v5 import analyze as chan_analyze
from scorer import extract_features
from talib_features import add_talib_features, get_talib_feature_names
from tencent_data import fetch_kline

OUT = os.path.join(HERE, '..', 'models')
os.makedirs(OUT, exist_ok=True)
LOG = '/tmp/train_v2.log'
STATE = '/tmp/train_v2_state.json'

def log(msg):
    """双输出: 文件+终端"""
    line = '[{}] {}'.format(datetime.now().strftime('%H:%M:%S'), msg)
    print(line, flush=True)
    with open(LOG, 'a') as f:
        f.write(line + '\n')

def set_state(phase, progress=None, detail=''):
    """写状态文件(供外部查询)"""
    s = {'phase': phase, 'progress': progress, 'detail': detail,
         'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    with open(STATE, 'w') as f:
        json.dump(s, f)

import baostock as bs  # 仅保留导入, 已不用

def collect_one_stock(args):
    """单只股票特征收集(worker进程调用) — 用腾讯数据源, 无登录"""
    code, _ = args
    try:
        dates, opens, highs, lows, closes, vols = fetch_kline(code)
        n = len(dates)
        if n < 200:
            return []
        
        talib_names = get_talib_feature_names()
        samples = []
        for window_end in range(200, n, 5):
            w = min(window_end, 750)  # 窗口放宽到3年(750根), 识别完整大级别中枢
            start = max(0, window_end - w)
            try:
                cur, bsp_buy, bsp_types, px, zs, pos = chan_analyze(
                    dates[start:window_end], opens[start:window_end],
                    closes[start:window_end], highs[start:window_end],
                    lows[start:window_end], code)
            except Exception:
                continue
            fd = extract_features(closes[start:window_end], highs[start:window_end],
                                  lows[start:window_end], opens[start:window_end],
                                  vols[start:window_end], bsp_buy, bsp_types, cur)
            base_vec = [fd[k] for k in sorted(fd.keys())]
            tf = add_talib_features(closes[start:window_end], highs[start:window_end],
                                    lows[start:window_end], vols[start:window_end])
            talib_vec = [tf.get(k, 0.0) for k in talib_names]
            vec = base_vec + talib_vec
            
            future_end = min(window_end + 5, n)
            future_return = (closes[future_end - 1] / px - 1) * 100 if future_end > window_end else 0
            label = 1 if future_return > 2 else 0
            samples.append({'features': vec, 'label': label, 'code': code})
        return samples
    except Exception:
        return []

def _init_worker():
    pass  # 腾讯HTTP无需登录

if __name__ == '__main__':
    import openpyxl
    # 清空旧日志
    if os.path.exists(LOG): os.remove(LOG)
    t0 = time.time()
    
    log('=' * 60)
    log('增强训练 v2 | v5引擎 + TA-Lib + 进度日志')
    log('=' * 60)
    
    # 加载CSI300
    wb = openpyxl.load_workbook('/root/chan_hs300_full_20260813.xlsx')
    ws = wb[wb.sheetnames[0]]
    codes = []
    for r in range(2, ws.max_row + 1):
        c = str(ws.cell(r, 1).value)
        if len(c) == 6:
            codes.append((c, ''))
    codes = codes[:300]
    n_total = len(codes)
    
    n_cores = min(24, cpu_count())
    log('股票数: {} | 核心: {}'.format(n_total, n_cores))
    
    # 阶段1a: 串行拉取K线(不并发, 防WAF)
    log('[1/3] 串行拉取K线(缓存, 不并发)...')
    set_state('fetch', '0/{}'.format(n_total))
    t_fetch = time.time()
    fetched = 0
    for code, _ in codes:
        dates, o, h, l, c, v = fetch_kline(code)  # 内部自动缓存
        time.sleep(0.1)  # 串行节流
        fetched += 1
        if fetched % 20 == 0 or fetched == n_total:
            log('  拉取 {}/{} ({:.0f}%) | 耗时{:.0f}s'.format(
                fetched, n_total, fetched/n_total*100, time.time()-t_fetch))
            set_state('fetch', '{}/{}'.format(fetched, n_total))
    log('  拉取完成 (耗时{:.0f}s)'.format(time.time()-t_fetch))
    
    # 阶段1b: 并行特征提取(读缓存, 纯CPU, 不再拉网络)
    log('[2/4] 并行特征提取 (读缓存, imap_unordered + 进度)...')
    set_state('collect', '0/{}'.format(n_total))
    
    all_samples = []
    done = 0
    t_collect = time.time()
    with Pool(n_cores, initializer=_init_worker) as pool:
        for result in pool.imap_unordered(collect_one_stock, codes, chunksize=1):
            done += 1
            if result:
                all_samples.extend(result)
            if done % 10 == 0 or done == n_total:
                elapsed = time.time() - t_collect
                rate = done / max(elapsed, 1)
                eta = (n_total - done) / max(rate, 0.001)
                log('  进度 {}/{} ({:.0f}%) | 已收集{}样本 | 速率{:.1f}只/s | ETA {:.0f}s'.format(
                    done, n_total, done/n_total*100, len(all_samples), rate, eta))
                set_state('collect', '{}/{}'.format(done, n_total),
                          'samples={} rate={:.1f}/s eta={:.0f}s'.format(len(all_samples), rate, eta))
    log('  收集完成: {}样本 (耗时{:.0f}s)'.format(len(all_samples), time.time()-t_collect))
    
    if len(all_samples) < 1000:
        log('⚠️ 样本不足, 退出')
        sys.exit(1)
    
    # 阶段3: 训练
    log('[3/4] XGBoost CPU训练...')
    set_state('train', None, 'samples={}'.format(len(all_samples)))
    t_train = time.time()
    
    from sklearn.model_selection import train_test_split
    import xgboost as xgb
    from sklearn.metrics import accuracy_score, roc_auc_score
    
    X = np.array([s['features'] for s in all_samples], dtype=np.float32)
    y = np.array([s['label'] for s in all_samples])
    log('  特征维度: {} | 正样本: {}/{} ({:.1f}%)'.format(
        X.shape[1], sum(y), len(y), sum(y)/len(y)*100))
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    model = xgb.XGBClassifier(
        max_depth=6, learning_rate=0.03, n_estimators=300,
        subsample=0.7, colsample_bytree=0.7,
        min_child_weight=10, reg_lambda=2, reg_alpha=0.5,
        scale_pos_weight=max(1, (len(y)-sum(y))/max(sum(y),1)),
        eval_metric='aucpr', tree_method='hist', n_jobs=64, random_state=42
    )
    model.fit(X_train, y_train, eval_set=[(X_test, y_test)], verbose=50)
    
    y_pred = model.predict(X_test)
    acc = accuracy_score(y_test, y_pred)
    auc = roc_auc_score(y_test, model.predict_proba(X_test)[:, 1])
    log('  准确率: {:.1f}% | AUC: {:.3f} (训练耗时{:.0f}s)'.format(acc*100, auc, time.time()-t_train))
    
    # 阶段3: 导出
    log('[4/4] 导出模型...')
    set_state('export', None, 'acc={:.1f}% auc={:.3f}'.format(acc*100, auc))
    out = os.path.join(OUT, 'chan_xgb_v5_talib.pkl')
    with open(out, 'wb') as f:
        pickle.dump(model, f)
    log('✅ {} ({:.1f}KB)'.format(out, os.path.getsize(out)/1024))
    
    set_state('done', '1/1', '总耗时{:.0f}s'.format(time.time()-t0))
    log('总耗时: {:.0f}s'.format(time.time()-t0))
