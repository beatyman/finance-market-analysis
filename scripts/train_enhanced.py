#!/usr/bin/env python3
"""
XGBoost增强训练器 — v5引擎 + TA-Lib因子 + 64核并行 + GPU

流程:
  1. 多进程并行收集300只CSI300的K线+BSP回放特征
  2. 特征 = 58维(缠论+量价) + 43维(TA-Lib) = 101维
  3. XGBoost GPU训练 (2080Ti)
  4. 导出模型
"""
import os, sys, time, pickle, numpy as np
from multiprocessing import Pool, cpu_count

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
sys.path.insert(0, os.path.join(HERE, '..', 'chanpy'))

from chan_engine_v5 import analyze as chan_analyze
from scorer import extract_features
from talib_features import add_talib_features, get_talib_feature_names

OUT = os.path.join(HERE, '..', 'models')
os.makedirs(OUT, exist_ok=True)

import baostock as bs

# === 单只股票特征收集 (供多进程调用) ===
def collect_one_stock(args):
    code, _ = args
    bs_local = baostock_login()  # 每个进程独立登录
    try:
        suffix = 'sh' if code.startswith('6') else 'sz'
        symbol = suffix + '.' + code
        rs = bs_local.query_history_k_data_plus(symbol,
            'date,open,high,low,close,volume',
            start_date='2023-01-01', end_date='2026-08-13',
            frequency='d', adjustflag='2')
        rows = []
        while rs.error_code == '0' and rs.next():
            rows.append(rs.get_row_data())
        if len(rows) < 200:
            return []
        
        dates = [r[0] for r in rows]
        opens = [float(r[1]) for r in rows]
        highs = [float(r[2]) for r in rows]
        lows = [float(r[3]) for r in rows]
        closes = [float(r[4]) for r in rows]
        vols = [float(r[5]) for r in rows]
        n = len(dates)
        
        # TA-Lib特征名(固定顺序)
        talib_names = get_talib_feature_names()
        
        samples = []
        for window_end in range(200, n, 5):
            w = min(window_end, 300)
            start = max(0, window_end - w)
            seg_dates = dates[start:window_end]
            seg_opens = opens[start:window_end]
            seg_closes = closes[start:window_end]
            seg_highs = highs[start:window_end]
            seg_lows = lows[start:window_end]
            seg_vols = vols[start:window_end]
            
            try:
                cur, bsp_buy, bsp_types, px, zs, pos = chan_analyze(
                    seg_dates, seg_opens, seg_closes, seg_highs, seg_lows, code)
            except Exception:
                continue
            
            # 缠论特征
            fd = extract_features(seg_closes, seg_highs, seg_lows, seg_opens,
                                  seg_vols, bsp_buy, bsp_types, cur)
            base_vec = [fd[k] for k in sorted(fd.keys())]
            
            # TA-Lib特征
            tf = add_talib_features(seg_closes, seg_highs, seg_lows, seg_vols)
            talib_vec = [tf.get(k, 0.0) for k in talib_names]
            
            # 合并特征
            vec = base_vec + talib_vec
            
            # 标签: 未来5日收益 > 2%
            future_end = min(window_end + 5, n)
            future_return = (closes[future_end - 1] / px - 1) * 100 if future_end > window_end else 0
            label = 1 if future_return > 2 else 0
            
            samples.append({'features': vec, 'label': label, 'code': code})
        
        return samples
    except Exception as e:
        return []

def baostock_login():
    lg = bs.login()
    return bs

def _init_worker():
    bs.login()

# === 主训练 ===
if __name__ == '__main__':
    import openpyxl
    t0 = time.time()
    
    # 加载CSI300代码
    wb = openpyxl.load_workbook('/root/chan_hs300_full_20260813.xlsx')
    ws = wb[wb.sheetnames[0]]
    codes = []
    for r in range(2, ws.max_row + 1):
        c = str(ws.cell(r, 1).value)
        if len(c) == 6:
            codes.append((c, ''))
    codes = codes[:300]
    
    n_cores = min(64, cpu_count())
    print('=' * 60)
    print('增强训练: {}只股票 | {}核心并行 | v5引擎 + TA-Lib'.format(len(codes), n_cores))
    print('=' * 60)
    
    # 多进程并行收集
    print('[1/3] 并行收集特征...')
    bs.login()  # 主进程登录(供子进程复用连接可能不行, 子进程各自登录)
    with Pool(n_cores, initializer=_init_worker) as pool:
        results = pool.map(collect_one_stock, codes, chunksize=2)
    
    samples = [s for r in results for s in r]
    print('  收集完成: {}个样本'.format(len(samples)))
    
    if len(samples) < 1000:
        print('⚠️ 样本不足')
        sys.exit(1)
    
    # 训练
    print('[2/3] XGBoost GPU训练...')
    from sklearn.model_selection import train_test_split
    import xgboost as xgb
    from sklearn.metrics import accuracy_score, roc_auc_score
    
    X = np.array([s['features'] for s in samples], dtype=np.float32)
    y = np.array([s['label'] for s in samples])
    
    print('  特征维度: {} | 正样本: {}/{} ({:.1f}%)'.format(
        X.shape[1], sum(y), len(y), sum(y)/len(y)*100))
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    
    model = xgb.XGBClassifier(
        max_depth=6, learning_rate=0.03, n_estimators=300,
        subsample=0.7, colsample_bytree=0.7,
        min_child_weight=10, reg_lambda=2, reg_alpha=0.5,
        scale_pos_weight=max(1, (len(y)-sum(y))/max(sum(y),1)),
        eval_metric='aucpr', tree_method='gpu_hist', random_state=42
    )
    model.fit(X_train, y_train, eval_set=[(X_test, y_test)], verbose=False)
    
    y_pred = model.predict(X_test)
    acc = accuracy_score(y_test, y_pred)
    auc = roc_auc_score(y_test, model.predict_proba(X_test)[:, 1])
    print('  准确率: {:.1f}% | AUC: {:.3f}'.format(acc*100, auc))
    
    # 特征重要性
    imp = model.feature_importances_
    top_idx = np.argsort(imp)[-15:][::-1]
    print('  Top15特征:')
    for i in top_idx:
        print('    feat{}: {:.4f}'.format(i, imp[i]))
    
    # 导出
    print('[3/3] 导出模型...')
    out = os.path.join(OUT, 'chan_xgb_v5_talib.pkl')
    with open(out, 'wb') as f:
        pickle.dump(model, f)
    print('✅ {} ({:.1f}KB)'.format(out, os.path.getsize(out)/1024))
    print('总耗时: {:.0f}s'.format(time.time()-t0))
    
    bs.logout()
