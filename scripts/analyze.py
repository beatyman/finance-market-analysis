import numpy as np
#!/usr/bin/env python3
"""
缠论多维度分析 v3.0 — chan.py + XGB评分 + 知识库确认
  用法: python3 analyze.py 002475           # A股
        python3 analyze.py hk00700           # 港股
        python3 analyze.py --scan --market hk  # 全市场扫描
"""
import os,sys,argparse,openpyxl,time
from datetime import datetime
HERE=os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0,HERE)

from data import fetch_kline_a,fetch_kline_hk,load_a_stocks,load_hk_stocks,fetch_hk_quotes,fetch_a_quotes,fetch_kline
from chan_engine import analyze as chan_analyze,get_bsp_label
from scorer import extract_features,score_from_features
from chan_kb import evaluate
from smc_insight import smc_analysis
from macro import load_macro,macro_signal
from futures_sentiment import get_futures_position,analyze_sentiment
from volume_sector import volume_analysis,get_stock_sector,sector_analysis
from sector_heat import sector_signal,get_sector_heat
from event_calendar import format_calendar
import pickle

MODEL_PATH=os.path.join(HERE,'..','models','chan_xgb_hk.pkl')
_xgb_model=None

def load_model():
    global _xgb_model
    if _xgb_model is None and os.path.exists(MODEL_PATH):
        try:
            _xgb_model=pickle.load(open(MODEL_PATH,'rb'))
            print('  📦 已加载训练模型 (XGBoost)')
        except:pass
    return _xgb_model

def predict_score(feats):
    """混合评分: 优先XGBoost模型，回退规则"""
    model=load_model()
    feat_len=len(feats)
    if model:
        vec=np.array([[feats[k] for k in sorted(feats.keys())]])
        # 尝试模型推理，维度不匹配则回退
        try:
            if model.n_features_in_==feat_len:
                proba=model.predict_proba(vec)[0,1]
                return int(proba*100)
        except:
            pass
    return score_from_features(feats)

def analyze_single(code_or_name,market='a'):
    """单股分析"""
    print('='*60)
    print('缠论多维度分析: %s'%code_or_name)
    print('='*60)
    
    # ── Event Calendar ──
    print()
    print(format_calendar())
    print()
    
    # ── Macro overview ──
    try:
        from macro import load_macro,macro_signal
        from futures_sentiment import get_futures_position,analyze_sentiment
        from futures_analysis import analyze_all_futures
        
        macro=load_macro();ms=macro_signal(macro)
        fp=get_futures_position();fs=analyze_sentiment(fp)
        fm=analyze_all_futures()
        
        print()
        print('🌍 宏观环境:')
        print('  📈 股指期货: %s'%fs['bias'])
        print('  💵 美元: DXY %.1f (%s)'%(macro.get('DXY',{}).get('value',0),'强势' if macro.get('DXY',{}).get('value',100)>101 else '中性'))
        print('  📊 美债10Y: %.2f%% (%s)'%(macro.get('US10Y',{}).get('value',0),'高利率' if macro.get('US10Y',{}).get('value',0)>4.5 else '中性'))
        print('  🇨🇳 USD/CNY: %.2f'%macro.get('USDCNY',{}).get('value',0))
        print('  🥇 商品: ',end='')
        comm=[] 
        for f in fm:
            if f:comm.append('%s %s %+.1f%%'%(f['name'].replace('COMEX',''),f['bsp'],f['chg_pct']))
        print(' | '.join(comm[:3]))
        print()
    except:pass
    
    print('─ 个股分析 ─')
    
    if market=='hk':
        data=fetch_kline_hk(code_or_name)
    else:
        data=fetch_kline_a(code_or_name)
    
    if not data:
        print('❌ K线数据获取失败')
        return
    dates,opens,closes,highs,lows,vols=data
    px=closes[-1]
    
    # Step 1: chan.py
    print('[1/3] chan.py 结构分析...')
    cur,bsp_buy,bsp_types,px,zs,pos=chan_analyze(dates,opens,closes,highs,lows,code_or_name)
    label=get_bsp_label(bsp_buy,bsp_types,pos)
    types_str='/'.join(bsp_types) if bsp_types else '-'
    print('  BSP: %s | 中枢: %s | 位置: %s | Bi:%d ZS:%d'%(label,zs or '无',pos,len(cur.bi_list) if cur else 0,len(cur.zs_list) if cur else 0))
    
    # Step 2: scoring
    print('[2/3] XGB 56维特征提取+打分...')
    fd=extract_features(closes,highs,lows,opens,vols,bsp_buy,bsp_types,cur);score=predict_score(fd)
    print('  评分: %d/100'%score)
    
    # Step 3: KB confirmation
    print('[3/3] 缠论知识库确认...')
    result=evaluate(code_or_name if market=='hk' else code_or_name,code_or_name,px,bsp_buy,bsp_types,pos,zs,score)
    
    print('\n📊 最终分析:')
    print('  标的: %s (¥%s)'%(result['name'],str(int(result['price']))))
    print('  方向: %s | 信号: %s | 评分: %d/100'%(result['direction'],result['signal'],result['score']))
    print('  风险: %s | 盈利预期: %s'%(result['risk'],result['profit'] or 'N/A'))
    
    # 阿娇标准: 年涨>100%的二买存疑
    if len(closes)>=120:
        ytd_chg=(closes[-1]/closes[-120]-1)*100
        if ytd_chg>100 and bsp_buy and '3' not in str(bsp_types):
            print('  ⚠️ 阿娇警告: 年涨%.0f%% + 二买 — 趋势中二买存疑，等三买确认'%ytd_chg)
    
    # 止损止盈
    if cur and cur.zs_list:
        z_last=cur.zs_list[-1];zl=float(z_last.low);zh=float(z_last.high)
        if bsp_buy:
            entry=int(zl);stop=int(zl*0.97);tp1=int(zh);tp2=int(zh*1.1)
            rr=abs(tp1-entry)/max(abs(entry-stop),1)
            print('  🎯 买入区: ¥%d~¥%d | 止损: ¥%d | TP1: ¥%d(+%d%%) | TP2: ¥%d(+%d%%) | R:R=%.1f:1'%(
                entry,int(zl*1.03),stop,tp1,int((tp1-px)/px*100),tp2,int((tp2-px)/px*100),rr))
    
    # 30分钟多级别确认
    try:
        import yfinance as yf
        sym_30m=code_or_name.replace('hk','')+'.HK' if market=='hk' else code_or_name+('.SS' if code_or_name.startswith('6') else '.SZ')
        df_30m=yf.download(sym_30m,period='5d',interval='30m',progress=False)
        if len(df_30m)>=50:
            def fv30(x):return float(x.item() if hasattr(x,'item') else x)
            c30=[fv30(x) for x in np.array(df_30m['Close']).ravel()][-200:]
            h30=[fv30(x) for x in np.array(df_30m['High']).ravel()][-200:]
            l30=[fv30(x) for x in np.array(df_30m['Low']).ravel()][-200:]
            o30=[fv30(x) for x in np.array(df_30m['Open']).ravel()][-200:]
            d30=[x.strftime('%Y-%m-%d %H:%M') for x in df_30m.index.tolist()][-200:]
            cur30,bsp_buy30,bsp_types30,px30,zs30,pos30=chan_analyze(d30,o30,c30,h30,l30,code_or_name+'_30m')
            label30=get_bsp_label(bsp_buy30,bsp_types30,pos30)
            if bsp_buy and bsp_buy30:
                confirm='✅ 日线+30m双买共振 — 高确信'
            elif bsp_buy and not bsp_buy30:
                confirm='🟡 日线买但30m未确认 — 等次级别'
            elif not bsp_buy and bsp_buy30:
                confirm='🟡 30m先于日线出买点 — 观察'
            else:
                confirm='—'
            print('  ⏱️ 30分钟: %s | %s | %s'%(label30,zs30 or '无中枢',confirm))
    except:pass
    
    smc=smc_analysis(code_or_name,result['name'],px,bsp_buy,zs,pos)
    print('  🧠 SMC视角: %s'%smc['verdict'])
    vol=volume_analysis([float(x) for x in closes],[float(x) for x in vols])
    print('  📊 量价: %s (vol_ratio=%.1fx)'%(vol['signal'],vol['vol_ratio']))
    sec=sector_analysis(code_or_name)
    print('  📈 板块: %s → %s'%(sec['sector'],sec['top_sector']))
    heat_data=get_sector_heat()
    sh=sector_signal(code_or_name,heat=heat_data)
    print('  🔥 板块热度: %s %+.1f%% (%d%%↑)'%(sh['signal'],sh['avg_chg'],sh['up_ratio']))
    if result['confirmations']:
        print('  ✅ 确认:')
        for c in result['confirmations']:print('    - '+c)
    
    return result

def scan_market(market='a',min_score=65):
    """全市场扫描 — 含量价/板块/阿娇过滤"""
    print('🔍 %s全市场扫描 (min_score≥%d)...'%({'a':'A股','hk':'港股'}[market],min_score))
    
    # 宏观缓存(只拉一次)
    try:
        macro=load_macro()
        fp=get_futures_position();fs=analyze_sentiment(fp)
        print('  📈 股指期货: %s | DXY: %.1f'%(fs['bias'],macro.get('DXY',{}).get('value',0)))
    except:pass
    
    # 板块热度缓存
    heat_data=get_sector_heat()
    print('  🔥 板块: %d个'%len(heat_data))
    
    if market=='a':
        stocks=load_a_stocks()
        print('  股票池: %d'%len(stocks))
        quotes=fetch_a_quotes(stocks)
        fetch_kline=fetch_kline_a
    else:
        stocks=load_hk_stocks()
        print('  股票池: %d'%len(stocks))
        quotes=fetch_hk_quotes(stocks)
        fetch_kline=fetch_kline_hk
    
    print('  行情: %d'%len(quotes))
    
    items=list(quotes.items())
    items.sort(key=lambda x:abs(x[1]['change_pct']),reverse=True)
    print('  全量分析: %d stocks...'%len(items))
    
    name_map={c:n for c,n in stocks}
    results=[]
    for idx,(code,q) in enumerate(items):
        if idx%30==0:print('    %d/%d, found %d'%(idx,len(items),len(results)),flush=True)
        name=name_map.get(code,'?')
        px=q['price'];chg=q['change_pct']
        
        data=fetch_kline(code)
        if not data:continue
        dates,opens,closes,highs,lows,vols=data
        
        cur,bsp_buy,bsp_types,px2,zs,pos=chan_analyze(dates,opens,closes,highs,lows,code)
        fd=extract_features(closes,highs,lows,opens,vols,bsp_buy,bsp_types,cur);score=predict_score(fd)
        
        if score<min_score:continue
        
        # 量价分析
        vol=volume_analysis([float(x) for x in closes],[float(x) for x in vols])
        vol_score=vol.get('score',50)
        
        # 板块热度
        sh=sector_signal(code,heat=heat_data)
        
        # 阿娇过滤: 年涨>100% + 非三买 → 降分
        ajiao_warn=''
        if len(closes)>=120:
            ytd_chg=(closes[-1]/closes[-120]-1)*100
            if ytd_chg>100 and bsp_buy and '3' not in str(bsp_types):
                ajiao_warn='⚠️二买存疑'
                score=max(0,score-15)
        
        # 综合评分: XGB 60% + 量价 20% + 板块 20%
        final_score=int(score*0.6+vol_score*0.2+sh.get('score',50)*0.2)
        
        label=get_bsp_label(bsp_buy,bsp_types,pos)
        results.append({
            'code':code,'name':name,'score':final_score,'xgb_score':score,
            'price':px,'change_pct':chg,
            'suggestion':label,
            'vol_signal':vol['signal'],'vol_ratio':vol['vol_ratio'],
            'sector':sh.get('sector','?'),'sector_signal':sh.get('signal','?'),
            'reason':'%s|%s|%s|%s|%s'%(pos,'/'.join(bsp_types) if bsp_types else '-',zs or '-',vol['signal'],ajiao_warn or ''),
            'ajiao':ajiao_warn,
            'time':datetime.now().strftime('%H:%M')
        })
    
    results.sort(key=lambda x:-x['score'])
    
    # Excel
    wb=openpyxl.Workbook();ws=wb.active;ws.title='缠论扫描'
    hdrs=['排名','代码','名称','综合分','XGB分','价格','涨跌幅','建议','量价','板块','板块信号','阿娇','理由','时间']
    from openpyxl.styles import Font,PatternFill
    hf=PatternFill(start_color='1a1a2e',end_color='1a1a2e',fill_type='solid')
    green=PatternFill(start_color='e8f5e9',end_color='e8f5e9',fill_type='solid')
    yellow=PatternFill(start_color='fff8e1',end_color='fff8e1',fill_type='solid')
    for c,h in enumerate(hdrs,1):ws.cell(1,c,h).fill=hf;ws.cell(1,c,h).font=Font(color='ffffff',bold=True)
    for i,r in enumerate(results):
        row=i+2;ws.cell(row,1,i+1)
        vals=[r['code'],r['name'],r['score'],r['xgb_score'],r['price'],r['change_pct'],
              r['suggestion'],r['vol_signal'],r['sector'],r['sector_signal'],r.get('ajiao',''),r['reason'],r['time']]
        for c,v in enumerate(vals,1):ws.cell(row,c+1,v)
        if r['score']>=75:
            for c in range(1,15):ws.cell(row,c).fill=green
        elif r.get('ajiao'):
            for c in range(1,15):ws.cell(row,c).fill=yellow
    for c,w in [('A',6),('B',10),('C',10),('D',8),('E',8),('F',8),('G',8),('H',16),('I',20),('J',12),('K',12),('L',12),('M',30),('N',8)]:
        ws.column_dimensions[c].width=w
    out=os.path.expanduser('~/chan_scan_%s.xlsx'%market)
    wb.save(out)
    
    print('\n🏆 Top 15:')
    for i,r in enumerate(results[:15]):
        ajiao=' %s'%r.get('ajiao','') if r.get('ajiao') else ''
        print(' %2d. %-10s %3d分 %+5.1f%% %s %s%s'%(i+1,r['name'],r['score'],r['change_pct'],r['suggestion'],r['vol_signal'],ajiao))
    print('Saved: %s | %d signals'%(out,len(results)))

if __name__=='__main__':
    p=argparse.ArgumentParser(description='缠论多维度分析')
    p.add_argument('code',nargs='?',help='股票代码/港股hk前缀')
    p.add_argument('--scan',action='store_true',help='全市场扫描')
    p.add_argument('--market',default='a',choices=['a','hk'],help='市场(a/hk)')
    p.add_argument('--min-score',type=int,default=60,help='最低评分')
    args=p.parse_args()
    
    if args.scan:
        scan_market(args.market,args.min_score)
    elif args.code:
        market='hk' if args.code.startswith('hk') else 'a'
        analyze_single(args.code,market)
    else:
        print('用法: python3 analyze.py <code> | --scan [--market a|hk]')
